import requests
from bs4 import BeautifulSoup
import openpyxl

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36 Edg/127.0.0.0"
}

total_bangumi = []
def save_bangumi_txt(all_titles,all_rating_nums,all_urls):

    for title, rating, url in zip(all_titles, all_rating_nums,all_urls):
        total_bangumi.append(f"{title}, {rating},{url}")
        
        #print(total_bangumi)

    with open("./bangumi.txt", "w", encoding="utf-8") as f:
        for bangumi in total_bangumi:
            #print(bangumi)
            f.write(bangumi + "\n")

def save_bangumi_xlsx(all_titles,all_rating_nums,all_urls):
    # 创建一个新的 Excel 工作簿
    wb = openpyxl.Workbook()

    # 获取默认的活动工作表
    sheet = wb.active

    # 设置表头（第一行）
    sheet['A1'] = '标题'
    sheet['B1'] = '评分'
    sheet['C1'] = '链接'

    # 写入数据到表格中
    for i in range(len(all_titles)):
        # 根据行数 i+2（从第二行开始），写入数据
        sheet.cell(row=i+2, column=1, value=all_titles[i])
        sheet.cell(row=i+2, column=2, value=all_rating_nums[i])
        sheet.cell(row=i+2, column=3, value=all_urls[i])

    # 保存 Excel 文件
    wb.save('bangumi.xlsx')

    # 提示用户保存成功
    print("Excel 文件 'bangumi.xlsx' 已成功创建并保存。")
            


def begin_scraping(headers):

    # 创建三个空列表来存储所有提取的标题和url和评分
    all_titles = []
    all_urls = []
    all_rating_nums = []    

    for start_num in range(0,100,25):
        response = requests.get(f"https://www.douban.com/doulist/116778197/?start={start_num}", headers=headers)


        if response.status_code == 200:
            print(f"进度%{start_num}.")
            #print(response.text[:500])  # 打印部分内容以验证请求结果

            bangumi_html = response.text
            soup = BeautifulSoup(bangumi_html,"html.parser")


            # 找到所有类名为 "title" 的 div 标签
            title_divs = soup.find_all('div', class_='title')

            # 遍历找到的 div 标签
            for div in title_divs:
                # 在每个 div 中找到 <a> 标签
                a_tag = div.find('a')
                if a_tag:
                    # 提取 <a> 标签内的文本内容并添加到列表中
                    title_text = a_tag.get_text(strip=True)
                    all_titles.append(title_text)
                    #提取url
                    url = a_tag.get('href')
                    all_urls.append(url)                

            all_rating = soup.find_all("span",attrs={"class":"rating_nums"})


            for rating_nums in all_rating:
                all_rating_nums.append(rating_nums.string)

        else:
            print("Request failed.")

    print("进度%100.")
    save_bangumi_xlsx(all_titles,all_rating_nums,all_urls)



begin_scraping(headers)
